import csv
import io
from calendar import monthrange
from datetime import date
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import PlainTextResponse, Response
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ...core.database import get_db
from ...crud.movement import LOW_STOCK_THRESHOLD_BOXES
from ...models.client import Client
from ...models.event import Event
from ...models.expense import Category, Expense
from ...models.finance import FixedAsset
from ...models.material import Material, MaterialPurchase
from ...models.movement import Movement
from ...models.product import Product
from ...models.sale import Sale, SaleItem
from ...models.user import User
from ..deps import get_current_user
from .pme import ar_aging as pme_ar_aging  # réutilise le calcul d'aging
from . import xlsx_i18n

router = APIRouter(prefix="/reports", tags=["reports"])

# Taux taxes QC — à factoriser dans config si elles changent
TPS_RATE = Decimal("0.05")
TVQ_RATE = Decimal("0.09975")


def _sale_tax_breakdown(sale: Sale) -> dict:
    """Calcule le détail TPS/TVQ d'une vente en se basant sur product.taxable de chaque ligne.
    Retourne {total_HT, taxable_subtotal, tps_collected, tvq_collected, total_TTC}.
    """
    total_HT = Decimal("0")
    taxable_subtotal = Decimal("0")
    for it in sale.items:
        sub = Decimal(it.subtotal)
        total_HT += sub
        if it.product is not None and getattr(it.product, "taxable", False):
            taxable_subtotal += sub
    tps = (taxable_subtotal * TPS_RATE).quantize(Decimal("0.01"))
    tvq = (taxable_subtotal * TVQ_RATE).quantize(Decimal("0.01"))
    return {
        "total_HT": total_HT.quantize(Decimal("0.01")),
        "taxable_subtotal": taxable_subtotal.quantize(Decimal("0.01")),
        "tps_collected": tps,
        "tvq_collected": tvq,
        "total_TTC": (total_HT + tps + tvq).quantize(Decimal("0.01")),
    }


def _event_cost_breakdown(e: Event) -> dict:
    return {
        "total_cost": (
            Decimal(e.registration_fee) + Decimal(e.transport_cost)
            + Decimal(e.other_costs) + Decimal(e.materials_cost)
        ).quantize(Decimal("0.01")),
        "profit": (
            Decimal(e.total_revenue)
            - Decimal(e.registration_fee) - Decimal(e.transport_cost)
            - Decimal(e.other_costs) - Decimal(e.materials_cost)
        ).quantize(Decimal("0.01")),
    }


def _month_range(year: int, month: int) -> tuple[date, date]:
    first = date(year, month, 1)
    last = date(year, month, monthrange(year, month)[1])
    return first, last


@router.get("/monthly")
def monthly_report(
    year: int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1, le=12),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict:
    """Aggregate monthly report — revenue, expenses, margin, top clients, low stock."""
    start, end = _month_range(year, month)

    # 1. Revenue (PAID sales only — what really hit the bank)
    revenue_paid = float(db.scalar(
        select(func.coalesce(func.sum(Sale.total_amount), 0))
        .where(Sale.status == "PAID", Sale.payment_date.between(start, end))
    ) or 0)

    # 2. Accounts receivable = DELIVERED but not PAID, regardless of date
    accounts_receivable = float(db.scalar(
        select(func.coalesce(func.sum(Sale.total_amount), 0))
        .where(Sale.status == "DELIVERED")
    ) or 0)

    # 3. Expenses
    expenses_total = float(db.scalar(
        select(func.coalesce(func.sum(Expense.amount), 0))
        .where(Expense.expense_date.between(start, end))
    ) or 0)

    expenses_by_cat = db.execute(
        select(Category.name, func.coalesce(func.sum(Expense.amount), 0).label("total"))
        .join(Expense, Expense.category_id == Category.id)
        .where(Expense.expense_date.between(start, end))
        .group_by(Category.name)
        .order_by(func.sum(Expense.amount).desc())
    ).all()
    expenses_by_category = [{"category": name, "total": float(total)} for name, total in expenses_by_cat]

    # 3b. Événements du mois — revenus encaissés sur place + coûts liés
    # (les ventes sur événements ne passent pas par la table Sale, elles sont saisies via Event)
    events_in_month = db.scalars(
        select(Event)
        .where(Event.start_date.between(start, end), Event.status != "CANCELLED")
    ).all()
    events_revenue = sum(float(e.total_revenue) for e in events_in_month)
    events_cost = sum(
        float(e.registration_fee) + float(e.transport_cost)
        + float(e.other_costs) + float(e.materials_cost)
        for e in events_in_month
    )
    # On fold dans les totaux globaux : événement = cash + dépenses réelles
    revenue_paid += events_revenue
    expenses_total += events_cost
    if events_cost > 0:
        expenses_by_category.append({"category": "Événements (festivals)", "total": events_cost})
        # Re-trier pour garder la cohérence
        expenses_by_category.sort(key=lambda x: x["total"], reverse=True)

    # 4. Net profit (revenue paid − expenses this month, événements inclus)
    net_profit = revenue_paid - expenses_total

    # 5. Sales by product (toutes non-annulées + sous-total PAID seulement)
    sales_by_product_rows = db.execute(
        select(
            Product.id, Product.name, Product.sku, Product.units_per_box, Product.unit_cost,
            func.coalesce(func.sum(SaleItem.quantity_boxes), 0).label("boxes"),
            func.coalesce(func.sum(SaleItem.subtotal), 0).label("revenue"),
        )
        .join(SaleItem, SaleItem.product_id == Product.id)
        .join(Sale, Sale.id == SaleItem.sale_id)
        .where(Sale.sale_date.between(start, end), Sale.status != "CANCELLED")
        .group_by(Product.id, Product.name, Product.sku, Product.units_per_box, Product.unit_cost)
        .order_by(func.sum(SaleItem.subtotal).desc())
    ).all()

    # 5b. Revenus encaissés (PAID) par produit
    paid_rows = db.execute(
        select(
            SaleItem.product_id,
            func.coalesce(func.sum(SaleItem.subtotal), 0).label("revenue_paid"),
            func.coalesce(func.sum(SaleItem.quantity_boxes), 0).label("boxes_paid"),
        )
        .join(Sale, Sale.id == SaleItem.sale_id)
        .where(Sale.sale_date.between(start, end), Sale.status == "PAID")
        .group_by(SaleItem.product_id)
    ).all()
    paid_by_pid = {r.product_id: (float(r.revenue_paid), int(r.boxes_paid)) for r in paid_rows}

    sales_by_product = []
    margin_by_product = []
    for row in sales_by_product_rows:
        boxes = int(row.boxes)
        revenue = float(row.revenue)
        cost = float(boxes * row.units_per_box * (row.unit_cost or 0))
        margin = revenue - cost
        margin_pct = (margin / revenue * 100) if revenue else 0

        # Marge encaissée (sur ventes PAID seulement)
        rev_paid, boxes_paid = paid_by_pid.get(row.id, (0.0, 0))
        cost_paid = float(boxes_paid * row.units_per_box * (row.unit_cost or 0))
        margin_paid = rev_paid - cost_paid
        margin_paid_pct = (margin_paid / rev_paid * 100) if rev_paid else 0

        sales_by_product.append({
            "product_id": row.id, "product_name": row.name, "product_sku": row.sku,
            "boxes_sold": boxes, "revenue": revenue,
        })
        margin_by_product.append({
            "product_id": row.id, "product_name": row.name,
            "revenue": revenue, "cost": cost, "margin": margin, "margin_pct": round(margin_pct, 1),
            "revenue_paid": rev_paid, "margin_paid": margin_paid,
            "margin_paid_pct": round(margin_paid_pct, 1),
        })

    # 6. Top clients (by revenue this month, excluding cancelled)
    top_clients_rows = db.execute(
        select(Client.id, Client.name, Client.type,
               func.coalesce(func.sum(Sale.total_amount), 0).label("total"))
        .join(Sale, Sale.client_id == Client.id)
        .where(Sale.sale_date.between(start, end), Sale.status != "CANCELLED")
        .group_by(Client.id, Client.name, Client.type)
        .order_by(func.sum(Sale.total_amount).desc())
        .limit(5)
    ).all()
    top_clients = [
        {"client_id": r.id, "client_name": r.name, "client_type": r.type, "total": float(r.total)}
        for r in top_clients_rows
    ]

    # 7. Inventory value + low-stock alerts
    by_pid = dict(db.execute(
        select(Movement.product_id, func.coalesce(func.sum(Movement.quantity_boxes), 0))
        .group_by(Movement.product_id)
    ).all())
    inventory_value = 0.0
    low_stock_alerts = []
    for p in db.scalars(select(Product)).all():
        boxes = int(by_pid.get(p.id, 0))
        if p.unit_cost is not None:
            inventory_value += boxes * p.units_per_box * float(p.unit_cost)
        if boxes < LOW_STOCK_THRESHOLD_BOXES:
            low_stock_alerts.append({
                "product_id": p.id, "product_name": p.name, "product_sku": p.sku,
                "stock_boxes": boxes,
            })

    return {
        "year": year, "month": month,
        "revenue_paid": round(revenue_paid, 2),
        "accounts_receivable": round(accounts_receivable, 2),
        "expenses_total": round(expenses_total, 2),
        "expenses_by_category": expenses_by_category,
        "net_profit": round(net_profit, 2),
        "sales_by_product": sales_by_product,
        "margin_by_product": margin_by_product,
        "top_clients": top_clients,
        "inventory_value": round(inventory_value, 2),
        "low_stock_alerts": low_stock_alerts,
        # Breakdown événements pour affichage dédié sur le dashboard
        "events_revenue": round(events_revenue, 2),
        "events_cost": round(events_cost, 2),
        "events_count": len(events_in_month),
    }


# ── CSV / XLSX exports ────────────────────────────────────────────────────────
# Headers comptables — réutilisés pour CSV + onglets XLSX
SALES_HEADERS = [
    "sale_id", "date", "client", "client_type", "status", "currency",
    "total_HT", "taxable_subtotal", "tps_collectee", "tvq_collectee", "total_TTC",
    "payment_date", "product_sku", "product_name", "product_taxable",
    "quantity_boxes", "unit_price", "subtotal",
]
EXPENSES_HEADERS = [
    "expense_id", "date", "account_code", "category", "expense_type",
    "amount_total", "amount_HT",
    "tps_paid", "tvq_paid", "taxes_total", "currency",
    "vendor", "vendor_tps_number", "vendor_tvq_number",
    "product", "description", "paid_by", "receipt_url",
    "is_recurring", "recurrence_frequency", "cca_class",
    "deductibility_pct", "deductible_amount",
]
EVENTS_HEADERS = [
    "event_id", "name", "location", "start_date", "end_date", "status",
    "total_revenue", "registration_fee", "transport_cost", "other_costs",
    "materials_cost", "total_cost", "profit", "units_sold",
    "material_label", "material_amount", "material_quantity", "material_unit",
    "material_id_catalog", "material_purchase_id", "notes",
]
MATERIAL_PURCHASES_HEADERS = [
    "purchase_id", "date", "material", "unit", "quantity",
    "total_cost", "unit_price", "vendor", "paid_by", "notes",
]


def _build_sales_rows(db: Session, start: Optional[date], end: Optional[date]) -> list[list]:
    stmt = select(Sale).order_by(Sale.sale_date.desc())
    if start and end:
        stmt = stmt.where(Sale.sale_date.between(start, end))
    rows: list[list] = []
    for sale in db.scalars(stmt).unique().all():
        tx = _sale_tax_breakdown(sale)
        common = [
            sale.id, sale.sale_date.isoformat(),
            sale.client.name if sale.client else "",
            sale.client.type if sale.client else "",
            sale.status, sale.currency,
            float(tx["total_HT"]), float(tx["taxable_subtotal"]),
            float(tx["tps_collected"]), float(tx["tvq_collected"]),
            float(tx["total_TTC"]),
            sale.payment_date.isoformat() if sale.payment_date else "",
        ]
        if not sale.items:
            rows.append(common + ["", "", "", "", "", ""])
            continue
        for it in sale.items:
            rows.append(common + [
                it.product.sku if it.product else "",
                it.product.name if it.product else "",
                "OUI" if (it.product and getattr(it.product, "taxable", False)) else "NON",
                it.quantity_boxes, float(it.unit_price), float(it.subtotal),
            ])
    return rows


def _build_expenses_rows(db: Session, start: Optional[date], end: Optional[date]) -> list[list]:
    stmt = select(Expense).order_by(Expense.expense_date.desc())
    if start and end:
        stmt = stmt.where(Expense.expense_date.between(start, end))
    rows: list[list] = []
    for ex in db.scalars(stmt).all():
        amount = Decimal(ex.amount)
        tps = Decimal(ex.tps_paid) if ex.tps_paid is not None else Decimal("0")
        tvq = Decimal(ex.tvq_paid) if ex.tvq_paid is not None else Decimal("0")
        amount_ht = amount - tps - tvq
        # Type = override de la dépense ou défaut catégorie
        exp_type = ex.expense_type or (ex.category.expense_type if ex.category else "OPEX")
        account_code = ex.category.account_code if ex.category else ""
        ded_pct = ex.deductibility_pct if ex.deductibility_pct is not None else 100
        deductible = float(amount_ht) * ded_pct / 100
        rows.append([
            ex.id, ex.expense_date.isoformat(),
            account_code or "",
            ex.category.name if ex.category else "",
            exp_type,
            float(amount), float(amount_ht),
            float(tps), float(tvq), float(tps + tvq),
            ex.currency,
            ex.vendor or "",
            ex.vendor_tps_number or "",
            ex.vendor_tvq_number or "",
            ex.product.name if ex.product else "",
            ex.description,
            ex.paid_by or "",
            ex.receipt_url or "",
            "OUI" if ex.is_recurring else "NON",
            ex.recurrence_frequency or "",
            ex.cca_class or "",
            ded_pct,
            round(deductible, 2),
        ])
    return rows


def _build_events_rows(db: Session, start: Optional[date], end: Optional[date]) -> list[list]:
    stmt = select(Event).order_by(Event.start_date.desc())
    if start and end:
        stmt = stmt.where(Event.start_date.between(start, end))
    rows: list[list] = []
    for ev in db.scalars(stmt).all():
        cb = _event_cost_breakdown(ev)
        base = [
            ev.id, ev.name, ev.location or "",
            ev.start_date.isoformat(),
            ev.end_date.isoformat() if ev.end_date else "",
            ev.status,
            float(ev.total_revenue),
            float(ev.registration_fee), float(ev.transport_cost),
            float(ev.other_costs), float(ev.materials_cost),
            float(cb["total_cost"]), float(cb["profit"]),
            ev.units_sold,
        ]
        breakdown = ev.materials_breakdown or []
        if not breakdown:
            rows.append(base + ["", "", "", "", "", "", ev.notes or ""])
        else:
            for item in breakdown:
                rows.append(base + [
                    item.get("label", ""),
                    float(item.get("amount") or 0),
                    item.get("quantity") if item.get("quantity") is not None else "",
                    item.get("unit") or "",
                    item.get("material_id") if item.get("material_id") is not None else "",
                    item.get("purchase_id") if item.get("purchase_id") is not None else "",
                    ev.notes or "",
                ])
    return rows


def _build_material_purchases_rows(db: Session, start: Optional[date], end: Optional[date]) -> list[list]:
    stmt = select(MaterialPurchase).order_by(MaterialPurchase.purchase_date.desc())
    if start and end:
        stmt = stmt.where(MaterialPurchase.purchase_date.between(start, end))
    rows: list[list] = []
    for p in db.scalars(stmt).unique().all():
        rows.append([
            p.id, p.purchase_date.isoformat(),
            p.material.name if p.material else "",
            p.material.unit if p.material else "",
            float(p.quantity),
            float(p.total_cost), float(p.unit_price),
            p.vendor or "", p.paid_by or "",
            p.notes or "",
        ])
    return rows


def _csv_response(headers: list[str], rows: list[list], filename: str) -> PlainTextResponse:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    w.writerows(rows)
    return PlainTextResponse(
        buf.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/exports/sales.csv", response_class=PlainTextResponse)
def export_sales_csv(
    year: Optional[int] = None,
    month: Optional[int] = None,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> PlainTextResponse:
    s, e = _month_range(year, month) if year and month else (None, None)
    return _csv_response(SALES_HEADERS, _build_sales_rows(db, s, e), "pinotte_sales.csv")


@router.get("/exports/expenses.csv", response_class=PlainTextResponse)
def export_expenses_csv(
    year: Optional[int] = None,
    month: Optional[int] = None,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> PlainTextResponse:
    s, e = _month_range(year, month) if year and month else (None, None)
    return _csv_response(EXPENSES_HEADERS, _build_expenses_rows(db, s, e), "pinotte_expenses.csv")


@router.get("/exports/events.csv", response_class=PlainTextResponse)
def export_events_csv(
    year: Optional[int] = None,
    month: Optional[int] = None,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> PlainTextResponse:
    s, e = _month_range(year, month) if year and month else (None, None)
    return _csv_response(EVENTS_HEADERS, _build_events_rows(db, s, e), "pinotte_events.csv")


@router.get("/exports/material_purchases.csv", response_class=PlainTextResponse)
def export_material_purchases_csv(
    year: Optional[int] = None,
    month: Optional[int] = None,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> PlainTextResponse:
    s, e = _month_range(year, month) if year and month else (None, None)
    return _csv_response(
        MATERIAL_PURCHASES_HEADERS,
        _build_material_purchases_rows(db, s, e),
        "pinotte_material_purchases.csv",
    )


@router.get("/exports/all.xlsx")
def export_all_xlsx(
    year: Optional[int] = None,
    month: Optional[int] = None,
    lang: str = "fr",
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Response:
    """Bundle XLSX comptable PCGR-friendly. 9 onglets :
    1. Résumé taxes (TPS/TVQ collectées vs payées vs nettes)
    2. État des résultats (income statement formaté : Revenus → COGS → Marge → OPEX → Net)
    3. Comptes à recevoir (AR aging : 0-30/31-60/61-90/90+)
    4. Ventes (avec colonnes taxes décomposées)
    5. Dépenses (avec account_code, type, déductibilité, CCA)
    6. Événements (festivals + breakdown matières)
    7. Achats matières premières
    8. Immobilisations (CCA + amortissement)
    9. Abonnements récurrents (vue annualisée)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    lang_norm = "en" if (lang or "").lower().startswith("en") else "fr"
    L = xlsx_i18n.get_labels(lang_norm)  # type: ignore[arg-type]

    s, e = _month_range(year, month) if year and month else (None, None)
    sales_rows = _build_sales_rows(db, s, e)
    expenses_rows = _build_expenses_rows(db, s, e)
    events_rows = _build_events_rows(db, s, e)
    purchases_rows = _build_material_purchases_rows(db, s, e)
    period_label = f"{year}-{month:02d}" if year and month else L["all_history"]

    # Patch "OUI/NON" → "YES/NO" en EN (sur expenses col 18) et idem pour product_taxable
    if lang_norm == "en":
        for row in expenses_rows:
            row[18] = "YES" if row[18] == "OUI" else "NO"
        for row in sales_rows:
            # product_taxable est à l'index 14
            if len(row) > 14:
                row[14] = "YES" if row[14] == "OUI" else "NO" if row[14] == "NON" else row[14]

    # Agrégats pour la feuille "Résumé taxes"
    # Ventes : on dédoublonne par sale_id pour ne pas multiplier les taxes par # de lignes
    seen_sales: set[int] = set()
    tps_collected_total = Decimal("0")
    tvq_collected_total = Decimal("0")
    revenue_ht_total = Decimal("0")
    for r in sales_rows:
        sid = r[0]
        if sid in seen_sales:
            continue
        seen_sales.add(sid)
        revenue_ht_total += Decimal(str(r[6]))      # total_HT
        tps_collected_total += Decimal(str(r[8]))   # tps_collectee
        tvq_collected_total += Decimal(str(r[9]))   # tvq_collectee
    # Dépenses : taxes payées agrégées (chaque dépense = 1 ligne)
    # Index dans EXPENSES_HEADERS : 6=amount_HT, 7=tps_paid, 8=tvq_paid
    tps_paid_total = sum((Decimal(str(r[7])) for r in expenses_rows), Decimal("0"))
    tvq_paid_total = sum((Decimal(str(r[8])) for r in expenses_rows), Decimal("0"))
    expenses_ht_total = sum((Decimal(str(r[6])) for r in expenses_rows), Decimal("0"))

    wb = Workbook()
    # Onglet par défaut = Ventes (sera après Résumé + P&L à la fin du build)
    ws_sales = wb.active
    ws_sales.title = L["sheet.sales"]

    def write_sheet(ws, headers: list[str], rows: list[list]):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="C5532E")  # paprika
            cell.alignment = Alignment(horizontal="left")
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A2"
        # Auto-width simple : longueur max sur les 50 premières lignes
        for col_idx, _h in enumerate(headers, start=1):
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, min(ws.max_row, 50) + 1)),
                default=10,
            )
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(max_len + 2, 10), 40)

    # Headers traduits par langue
    sales_h = xlsx_i18n.headers(xlsx_i18n.SALES_HEADER_KEYS, lang_norm)  # type: ignore[arg-type]
    exp_h = xlsx_i18n.headers(xlsx_i18n.EXPENSES_HEADER_KEYS, lang_norm)  # type: ignore[arg-type]
    ev_h = xlsx_i18n.headers(xlsx_i18n.EVENTS_HEADER_KEYS, lang_norm)  # type: ignore[arg-type]
    mp_h = xlsx_i18n.headers(xlsx_i18n.MATERIAL_PURCHASES_HEADER_KEYS, lang_norm)  # type: ignore[arg-type]

    write_sheet(ws_sales, sales_h, sales_rows)
    write_sheet(wb.create_sheet(L["sheet.expenses"]), exp_h, expenses_rows)
    write_sheet(wb.create_sheet(L["sheet.events"]), ev_h, events_rows)
    write_sheet(wb.create_sheet(L["sheet.material_purchases"]), mp_h, purchases_rows)

    # ── État des résultats (income statement PCGR) ─────────────
    # Sépare Revenus / COGS / Marge brute / Charges d'exploitation / Bénéfice net
    revenue_sales = sum((Decimal(str(r[6])) for r in sales_rows if r[0] not in set()), Decimal("0"))
    # On dédoublonne par sale_id sur total_HT (col 6 dans SALES_HEADERS)
    seen_ids: set[int] = set()
    revenue_sales = Decimal("0")
    for r in sales_rows:
        if r[0] in seen_ids:
            continue
        seen_ids.add(r[0])
        revenue_sales += Decimal(str(r[6]))
    # Revenus événements = total_revenue par event_id (col 6 dans EVENTS_HEADERS)
    seen_events: set[int] = set()
    revenue_events = Decimal("0")
    for r in events_rows:
        if r[0] in seen_events:
            continue
        seen_events.add(r[0])
        revenue_events += Decimal(str(r[6]))

    # Décompose les dépenses par type
    cogs_total = Decimal("0")
    opex_total = Decimal("0")
    capex_total = Decimal("0")
    opex_by_account: dict[str, Decimal] = {}
    cogs_by_account: dict[str, Decimal] = {}
    for r in expenses_rows:
        amount_ht = Decimal(str(r[6]))  # amount_HT
        exp_type = r[4]
        cat = r[3]
        code = r[2] or ""
        label = f"{code} {cat}".strip() if code else cat
        if exp_type == "COGS":
            cogs_total += amount_ht
            cogs_by_account[label] = cogs_by_account.get(label, Decimal("0")) + amount_ht
        elif exp_type == "CAPEX":
            capex_total += amount_ht
        else:  # OPEX par défaut
            opex_total += amount_ht
            opex_by_account[label] = opex_by_account.get(label, Decimal("0")) + amount_ht

    total_revenue = revenue_sales + revenue_events
    gross_margin = total_revenue - cogs_total
    gm_pct = float(gross_margin / total_revenue * 100) if total_revenue > 0 else 0
    operating_profit = gross_margin - opex_total
    op_pct = float(operating_profit / total_revenue * 100) if total_revenue > 0 else 0

    ws_pl = wb.create_sheet(L["sheet.income_statement"], 1)  # juste après Résumé taxes
    ws_pl["A1"] = f"{L['pl.title']} — {period_label}"
    ws_pl["A1"].font = Font(bold=True, size=14, color="C5532E")
    ws_pl["A2"] = L["pl.subtitle"]
    ws_pl["A2"].font = Font(italic=True, color="666666")

    def pl_row(label: str, amount: Optional[Decimal] = None, *, bold=False, indent=0, top_border=False):
        row_idx = ws_pl.max_row + 1
        cell_a = ws_pl.cell(row=row_idx, column=1, value=("  " * indent) + label)
        cell_b = ws_pl.cell(row=row_idx, column=2, value=float(amount) if amount is not None else None)
        if bold:
            cell_a.font = Font(bold=True)
            cell_b.font = Font(bold=True)
        if top_border:
            from openpyxl.styles import Border, Side
            border = Border(top=Side(style="thin"))
            cell_a.border = border
            cell_b.border = border

    pl_row("")
    pl_row(L["pl.section_revenue"], bold=True)
    pl_row(L["pl.net_sales"], revenue_sales, indent=1)
    pl_row(L["pl.events_revenue"], revenue_events, indent=1)
    pl_row(L["pl.total_revenue"], total_revenue, bold=True, indent=1, top_border=True)
    pl_row("")
    pl_row(L["pl.section_cogs"], bold=True)
    for k, v in sorted(cogs_by_account.items(), key=lambda x: x[1], reverse=True):
        pl_row(k, v, indent=1)
    if not cogs_by_account:
        pl_row(L["pl.cogs_empty"], indent=1)
    pl_row(L["pl.total_cogs"], cogs_total, bold=True, indent=1, top_border=True)
    pl_row("")
    pl_row(L["pl.gross_margin"], gross_margin, bold=True)
    pl_row(f"  ({gm_pct:.1f}{L['pl.pct_of_revenue']})" if total_revenue > 0 else f"  {L['pl.no_revenue_note']}", indent=1)
    pl_row("")
    pl_row(L["pl.section_opex"], bold=True)
    for k, v in sorted(opex_by_account.items(), key=lambda x: x[1], reverse=True):
        pl_row(k, v, indent=1)
    pl_row(L["pl.total_opex"], opex_total, bold=True, indent=1, top_border=True)
    pl_row("")
    pl_row(L["pl.operating_profit"], operating_profit, bold=True, top_border=True)
    pl_row(f"  ({op_pct:.1f}{L['pl.pct_of_revenue']})" if total_revenue > 0 else "", indent=1)
    pl_row("")
    if capex_total > 0:
        pl_row(L["pl.capex_note"], capex_total, indent=0)
        pl_row(f"  {L['pl.capex_explain']}", indent=1)
    ws_pl.column_dimensions["A"].width = 50
    ws_pl.column_dimensions["B"].width = 18

    # ── Comptes à recevoir (AR aging) ──────────────────────────
    ws_ar = wb.create_sheet(L["sheet.ar_aging"])
    aging = pme_ar_aging(as_of=None, db=db)  # type: ignore[arg-type]
    ws_ar.append(xlsx_i18n.headers(xlsx_i18n.AR_AGING_HEADER_KEYS, lang_norm))  # type: ignore[arg-type]
    for cell in ws_ar[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C5532E")
    for c in aging.by_client:
        ws_ar.append([
            c.client_name, c.client_type, c.invoice_count,
            float(c.days_0_30), float(c.days_31_60), float(c.days_61_90), float(c.days_90_plus),
            float(c.total),
        ])
    # Ligne TOTAL en bas
    from openpyxl.styles import Border, Side
    tot_row_idx = ws_ar.max_row + 1
    ws_ar.append([
        L["total"], "", "",
        float(aging.totals.days_0_30), float(aging.totals.days_31_60),
        float(aging.totals.days_61_90), float(aging.totals.days_90_plus),
        float(aging.totals.total),
    ])
    for cell in ws_ar[tot_row_idx]:
        cell.font = Font(bold=True)
        cell.border = Border(top=Side(style="medium"))
    # DSO en ligne info
    if aging.dso_days is not None:
        ws_ar.append([])
        ws_ar.append([f"{L['ar.dso_label']} : {aging.dso_days} {L['ar.dso_unit']}"])
        ws_ar.cell(row=ws_ar.max_row, column=1).font = Font(italic=True, color="666666")
    ws_ar.column_dimensions["A"].width = 28
    for col in ["B", "C", "D", "E", "F", "G", "H"]:
        ws_ar.column_dimensions[col].width = 14
    ws_ar.freeze_panes = "A2"

    # ── Immobilisations + amortissement CCA ────────────────────
    fixed_assets = db.scalars(select(FixedAsset).order_by(FixedAsset.purchase_date.desc())).all()
    ws_fa = wb.create_sheet(L["sheet.fixed_assets"])
    fa_headers = xlsx_i18n.headers(xlsx_i18n.FIXED_ASSETS_HEADER_KEYS, lang_norm)  # type: ignore[arg-type]
    write_sheet(ws_fa, fa_headers, [
        [
            a.name, a.purchase_date.isoformat(),
            float(a.cost), a.cca_class, float(a.cca_rate_pct),
            float(a.accumulated_depreciation),
            float(Decimal(a.cost) - Decimal(a.accumulated_depreciation)),
            float((Decimal(a.cost) - Decimal(a.accumulated_depreciation)) * Decimal(a.cca_rate_pct) / 100),
            a.disposal_date.isoformat() if a.disposal_date else "",
        ]
        for a in fixed_assets
    ])

    # ── Abonnements récurrents (vue annualisée) ────────────────
    sub_rows: list[list] = []
    seen_sub_keys: set[str] = set()
    # is_recurring est traduit ("OUI"/"YES") selon lang_norm — on accepte les deux
    recurring_yes = {"OUI", "YES"}
    for r in expenses_rows:
        if r[18] not in recurring_yes:
            continue
        key = f"{r[11]}|{r[15]}|{r[5]}"  # vendor + description + amount_total
        if key in seen_sub_keys:
            continue
        seen_sub_keys.add(key)
        amount = Decimal(str(r[5]))
        freq = r[19] or "monthly"
        mult = {"monthly": 12, "quarterly": 4, "yearly": 1}.get(freq, 12)
        sub_rows.append([
            r[11] or "—", r[15], freq, float(amount), float(amount * mult),
        ])
    ws_sub = wb.create_sheet(L["sheet.subscriptions"])
    sub_headers = xlsx_i18n.headers(xlsx_i18n.SUBSCRIPTIONS_HEADER_KEYS, lang_norm)  # type: ignore[arg-type]
    write_sheet(ws_sub, sub_headers, sub_rows)
    if sub_rows:
        total_annual = sum(r[4] for r in sub_rows)
        ws_sub.append([])
        ws_sub.append([L["sub.total_annualized"], "", "", "", total_annual])
        ws_sub.cell(row=ws_sub.max_row, column=1).font = Font(bold=True)
        ws_sub.cell(row=ws_sub.max_row, column=5).font = Font(bold=True)

    # Feuille "Résumé taxes" — onglet 0 (positionné après création)
    ws_tax = wb.create_sheet(L["sheet.tax_summary"], 0)
    ws_tax["A1"] = L["tax.title"]
    ws_tax["A1"].font = Font(bold=True, size=14, color="C5532E")
    ws_tax["A2"] = f"{L['tax.period']} {period_label}"
    ws_tax["A2"].font = Font(italic=True, color="666666")
    rows_summary = [
        [],
        [L["tax.section_collected"]],
        [L["tax.revenue_ht"], float(revenue_ht_total)],
        [L["tax.tps_collected"], float(tps_collected_total)],
        [L["tax.tvq_collected"], float(tvq_collected_total)],
        [L["tax.total_collected"], float(tps_collected_total + tvq_collected_total)],
        [],
        [L["tax.section_paid"]],
        [L["tax.expenses_ht"], float(expenses_ht_total)],
        [L["tax.tps_paid"], float(tps_paid_total)],
        [L["tax.tvq_paid"], float(tvq_paid_total)],
        [L["tax.total_recoverable"], float(tps_paid_total + tvq_paid_total)],
        [],
        [L["tax.section_net"]],
        [L["tax.net_tps"], float(tps_collected_total - tps_paid_total)],
        [L["tax.net_tvq"], float(tvq_collected_total - tvq_paid_total)],
        [L["tax.net_total"], float(
            (tps_collected_total + tvq_collected_total)
            - (tps_paid_total + tvq_paid_total)
        )],
    ]
    for r in rows_summary:
        ws_tax.append(r)
    # Mise en gras des titres de section
    for row_idx in (4, 10, 16):
        c = ws_tax.cell(row=row_idx, column=1)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="C5532E")
    ws_tax.column_dimensions["A"].width = 42
    ws_tax.column_dimensions["B"].width = 18

    # Sauvegarde dans un buffer mémoire
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"pinotte_export_{period_label}_{lang_norm}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
